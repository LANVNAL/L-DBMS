# -*- coding: utf-8 -*

from openpyxl import *
import  os
import re
from index import *
from prettytable import PrettyTable
import hashlib
using_dbname = ''
using_db = Workbook()
db_path = 'data/'
#view_path = 'view/'
user = ''



def welcome():
    """
    欢迎界面/字符画
    :return:
    """
    print("""
          ##############################################
          
                    https://github.com/LANVNAL 
              _          _____  ____  __  __  _____ 
             | |        |  __ \|  _ \|  \/  |/ ____|
             | |  ______| |  | | |_) | \  / | (___  
             | | |______| |  | |  _ <| |\/| |\___ \ 
             | |____    | |__| | |_) | |  | |____) |
             |______|   |_____/|____/|_|  |_|_____/ 
                                                    
                    -> exit:退出 help:语法帮助 <-

          ##############################################
          """)


def help():
    """
    打印帮助信息
    :return:
    """
    print("""
    1.创建表：create database dbname
    2.创建数据库：create table tbname (id int PK null,user char[10] )
    3.删除：DELETE FROM table_nmae WHERE column_name = 'Value'
    4.更新：UPDATE table_name SET column1=value1,column2=value2,... WHERE some_column=some_value;
    5.插入：INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
    6.查询：select a,b from table where c=x,d=x （与）
           select a,b from table where c=x|d=x（或）
           select a,b from table where c>x,d<x
           支持like，in，支持子查询
    7.权限：grant/revoke select on test_tb for testuser
    8.索引：creat view view_name as select xx from xx
    9.显示信息：help table/view/index
    """)


def get_command():
    """
    从控制台获取命令
    :return: None
    """
    command = raw_input("[👉]> ") if not using_dbname else raw_input("[{}🚩]> ".format(using_dbname))
    #hcommand = command.lower()
    #print command
    return command.strip()



def use_db(dbname):
    global using_dbname
    global using_db
    if check_permission(user, dbname, 'use'):
        using_dbname = dbname
        using_db = load_workbook(db_path+dbname+'.xlsx')
        print("Database changed.")

def show_db():
    print "All database:"
    dbs = os.listdir(db_path)   #第二种方法，从保存数据库信息的库中查询
    for db in dbs:
        if '.DS' not in db and db != 'index':
            print "[*] " + db[:-5]

def creat_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = Workbook()
    database.save(dbpath)
    create_tb_in_tbinfo(dbname)
    print(u"数据库创建操作执行成功")

def Initialization():
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists("data/table_information.xlsx"):
        Workbook().save("data/table_information.xlsx")
    if os.path.exists("data/system.xlsx"):
        print "Initializating......"
    else:
        creat_db('system')
    db = load_workbook("data/system.xlsx")
    permission_tb_col = ['database char[50] pk unique','select char','insert char','delete char','update char']
    creat_table('permission', db, 'system',permission_tb_col)

def create_tb_in_tbinfo(dbname):    #在table_infomation中创建数据库对应的表，
    db = load_workbook("data/table_information.xlsx")
    table = db.create_sheet(dbname)
    columns_name = ['table','column_name', 'type', 'null', 'unique', 'primary_key', 'foreign_key']
    for i in range(len(columns_name)):
        table.cell(row=1,column=i+1).value = columns_name[i]
    if db.worksheets[0].title == 'Sheet':
        del db['Sheet']
    db.save("data/table_information.xlsx")


#create table tbname (id int PK null,user char[10] )
def creat_table(table_name,current_database,current_dbname,columns_list):
    # create table
    if table_name not in current_database.sheetnames:
        table = current_database.create_sheet(table_name)
    else:
        print (u"数据表已存在,请重新输入.")
        return
    if current_database.worksheets[0].title == 'Sheet':
        del current_database['Sheet']
    #表创建完成，开始创建列
    length = len(columns_list)
    #print length
    tbinfo = load_workbook("data/table_information.xlsx")
    tbinfo_tb = tbinfo[current_dbname]
    tbinfo_rows = tbinfo_tb.max_row
    column_names = []
    for i in range(length):             #将字段的属性写到table_information库中
        column = columns_list[i].split(' ')
        tbinfo_tb.cell(row=tbinfo_rows+1+i,column=1).value = table_name
        tbinfo_tb.cell(row=tbinfo_rows+1+i, column=2).value = column[0]
        tbinfo_tb.cell(row=tbinfo_rows+1+i, column=3).value = column[1]
        for key in column[2:]:
            if key == 'null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '1'
            elif key == 'not_null':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=4).value = '0'
            elif key == 'unique':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=5).value = '1'
            elif key == 'pk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=6).value = '1'
            elif key == 'fk':
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=7).value = '1'
        column_names.append(column[0])
        for j in range(1, 8):
            if tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value is None:
                tbinfo_tb.cell(row=tbinfo_rows + 1 + i, column=j).value = 'NULL'
    tbinfo.save("data/table_information.xlsx")
    for i in range(length):
        table.cell(row=1,column=i+1).value = column_names[i]  #表第一行是列名
    current_dbname = db_path + current_dbname + '.xlsx'
    current_database.save(current_dbname)
    print (u"数据表创建完成。")


def insert(table_name, current_database, current_dbname, columns_list):
    if not check_Constraint(columns_list,table_name):    #columns应为[dict]
        print "Constraint Error"
        return False
    table = current_database[table_name]
    for columns in columns_list:
        table_rows = table.max_row
        table_columns = table.max_column
        length = len(columns)
        # print length
        for i in range(length):
            column = re.search('\((.*?)\)', columns[i], re.S).group(1)
            column_list = column.split(',')
            chk_len = len(column_list)
            if chk_len != table_columns:
                print ('插入失败，请检查输入的数据数量是否与列数量对应。')
                return

            else:
                for j in range(chk_len):
                    table.cell(row=table_rows + i + 1, column=j + 1).value = column_list[j]
                current_dbname = db_path + current_dbname + '.xlsx'
                current_database.save(current_dbname)
                print ("数据插入完成。")

#DELETE FROM table_nmae WHERE column_name = 'Value'
def delect(table_name,current_database,current_dbname,columns_list):  #columns_list={'name1':'value1','name2':'value2'}

    table = current_database[table_name]
    table_rows = table.max_row  #行
    table_columns = table.max_column    #列
    length = len(columns_list)
    delect_row_num = [x for x in range(2,table_rows+1)]
    columns_name=[]
    for cell in list(table.rows)[0]:
        columns_name.append(cell.value)
    for key in columns_list:
        flag = 0
        for i in range(len(columns_name)):    #判断colmuns_list 是否有 not in colmus中的
            if columns_name[i] == key:
                flag = 1
        if flag == 0:   #输入的列名不存在
            print("Unknown column '{}' in 'where clause'".format(key))
            return
    for key in columns_list:
        column_num = columns_name.index(key)
        for i in delect_row_num[::-1]:  #倒着来
            if table.cell(row=i, column=column_num+1).value != columns_list[key]:
                delect_row_num.remove(i)
    if len(delect_row_num) > 0:
        for i in delect_row_num[::-1]:
            #print i,table_rows
            table.delete_rows(int(i))
    else:
        print("find 0 to delect.")
    current_database.save(db_path + current_dbname + '.xlsx')
    print("删除完成，影响{}行".format(len(delect_row_num)))

#UPDATE table_name SET column1=value1,column2=value2,... WHERE some_column=some_value;
def update(table_name,current_database,current_dbname,columns_list,update_columns_list):
    if not check_Constraint(update_columns_list,table_name):    #columns应为dict
        print "Constraint Error"
        return False
    table = current_database[table_name]
    table_rows = table.max_row  # 行
    table_columns = table.max_column  # 列
    length = len(columns_list)
    update_row_num = [x for x in range(2,table_rows+1)]
    columns_name = []
    for cell in list(table.rows)[0]:
        columns_name.append(cell.value)
    for key in columns_list:
        flag = 0
        for i in range(len(columns_name)):  # 判断colmuns_list 是否有 not in colmus中的
            if columns_name[i] == key:
                flag = 1
        if flag == 0:  # 输入的列名不存在
            print("Unknown column '{}' in 'where clause'".format(key))
            return
    for key in columns_list:
        column_num = columns_name.index(key)
        for i in update_row_num[::-1]:  #倒着来
            if table.cell(row=i, column=column_num+1).value != columns_list[key]:
                update_row_num.remove(i)
    if len(update_row_num) > 0:
        for i in update_row_num[::-1]:
            for j in range(1,table_columns+1):
                clu_name = table.cell(row=1, column=j).value
                table.cell(row=i, column=j).value = update_columns_list[clu_name]
    else:
        print("find 0 to update.")
    current_database.save(db_path + current_dbname + '.xlsx')
    print("更新完成，影响{}行".format(len(update_row_num)))

def select_index(a):
    pos = BPTree_search(a)

def update_index(table_name,column_name):
    index(using_db,table_name,column_name)

def index(current_database,table_name,column_name):
    table = current_database[table_name]
    table_columns = table.max_column
    table_rows = table.max_row
    column_num = 0
    column_value = []
    column_position = []
    for i in range(1,table_columns+1):
        if table.cell(row=1,column=i).value == column_name:
            column_num = i
    if column_num == 0:
        print "no this column"
        return
    else:
        for i in range(2, table_rows+1):
            column_value.append(str(table.cell(row=i,column=column_num).value))
            column_position.append('<{},{}>'.format(i,column_num))
    column_value.sort()
    for i in range(len(column_value)):
        tmp = [column_value[i],column_position[i]]   #like [1,aaa|<2,1>]
        column_value[i] = tuple(tmp)    #like [(1,aaa|<2,1>)]
    #print column_value[0]
    bt = test_BPTree(column_value)
    indexname = table_name + '|' +column_name
    save_index(str(bt), indexname)

def save_index(bt,indexname):
    line = re.findall(r'\[.*?\]', bt)
    for i in range(len(line)):
        line[i] = line[i][2:-2].replace('|', '')
    file = open('data/index/' + indexname,'w')
    for i in range(len(line)):
        file.writelines(line[i] + '\n')
    file.close()


#select a,b from table where c=x,d=x
def select(columns,table_name,limit={},predicate='and', symbol='=', tag=''):    #{'c':'x','d':'x'}
    if using_dbname == '':
        print "please choose databse!"
        return
    table = using_db[table_name]
    #print columns
    if columns == '*' and len(limit) == 0:
        columns_name = list(iter_rows(table))[0]
        table_print = PrettyTable(columns_name)
        for i in range(1,len(list(iter_rows(table)))):
            table_print.add_row(list(iter_rows(table))[i])
        table_print.reversesort = True
        if tag == 'view':
            print table_print
            return list(iter_rows(table))   #view
        if tag == 'insert':
            return list(iter_rows(table))
        else:
            print(table_print)
    else:
        sel_cols = columns.split(',')   #*的情况
        rows_list = list(iter_rows(table))  #所有的行
        cols = rows_list[0]
        col_pos = []
        limit_pos = []
        print_row = []
        limit_cols = list(limit)
        symbol = '==' if symbol == '=' else symbol
        if columns[0] != '*':
            for i in range(len(sel_cols)):
                col_pos.append(cols.index(sel_cols[i])) #要查的列的列号
        else:
            sel_cols = list(iter_rows(table))[0]
            col_pos = range(len(cols))
        for i in range(len(limit)):
            limit_pos.append(cols.index(limit_cols[i])) #where的列
        for i in range(1, len(rows_list)):
            match = 0
            if predicate == 'in':
                match_list = limit[limit_cols[0]]
                for j in len(match_list):
                    if rows_list[i][limit_pos[0]] == match_list[j]:
                        print_row.append(i)
            if predicate == 'like':
                like_word = re.findall('(.*)\%',limit[limit_cols[0]])
                if like_word in rows_list[i][limit_pos[0]]:
                    print_row.append(i)
            else:
                for j in range(len(limit_pos)): #通过eval实现比较运算
                    if eval("'" + rows_list[i][limit_pos[j]] + "'" + symbol + "'" + limit[limit_cols[j]] + "'"):
                        match += 1
                if predicate == None:
                    print_row.append(i)
                if predicate == 'and' and match == len(limit_pos):  #and时要全部匹配
                    print_row.append(i)     #符合条件的行号
                if predicate == 'or' and match > 0: #or时至少一个匹配
                    print_row.append(i)

        table_print = PrettyTable(sel_cols)
        for i in range(len(print_row)):
            add_rows = []
            for x in col_pos:
                add_rows.append(rows_list[print_row[i]][x])
            table_print.add_row(add_rows)
        table_print.reversesort = True
        if tag == 'view':
            return table_print
        elif tag == 'insert':
            return table_print
        elif tag == 'nesting':
            tmpdb = using_db
            table = tmpdb['tmp']
            for i in range(len(sel_cols)):
                table.cell(row=0,column=i+1).value = sel_cols[i]
            for i in range(len(print_row)):
                add_rows = []
                for x in col_pos:
                    add_rows.append(rows_list[print_row[i]][x])
                for j in range(len(add_rows)):
                    table.cell(row=i+2,column=j+1).value = add_rows[j]
            tmpdb.save("data/" + using_dbname + ".xlsx")

        else:
            #table_print.reversesort = True
            print(table_print)

def nested_select():
    return

#creat view view_name as select xx from xx
def view(viewname,sql):
    """
    file = view_path + viewname
    view = query(sql,'view')
    f = open(file, "w")
    f.write(str(view))
    f.close()
    print "Success"
    """
    db = load_workbook("data/view.xlsx")
    if viewname not in db.sheetnames:
        table = db.create_sheet(viewname)
    else:
        print ("view is exist.")
        return
    if db.worksheets[0].title == 'Sheet':
        del db['Sheet']
    sql_table = db['sql']
    maxrow = sql_table.max_row  #在sql表中存view名和对应的sql语句
    sql_table.cell(row=maxrow + 1, column = 1).value = viewname
    sql_table.cell(row=maxrow + 1, column = 2).value = sql
    table = db[viewname]
    views = query(sql, 'view')
    for i in range(len(views)):
        for j in range(len(views[i])):
            table.cell(row=i+1, column=j+1).value = views[i][j]
    db.save("data/view.xlsx")



#grant select on test_tb for testuser
def set_permission(user,database,action):
    db = load_workbook("data/system.xlsx")
    table = db['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database) + 2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action) + 1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        print "user have this permission"
    else:
        table.cell(row=row, column=col).value = table.cell(row=row, column=col).value + ',' + user
        db.save("data/system.xlsx")

#revoke select on test_tb for testuser
def del_permission(user,database,action):
    db = load_workbook("data/system.xlsx")
    table = db['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database) + 2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action) + 1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        if allow_user.index(user) == 0:
            table.cell(row=row, column=col).value = table.cell(row=row, column=col).value.replace(user, '')
        else:
            table.cell(row=row, column=col).value = table.cell(row=row, column=col).value.replace(',' + user, '')
        db.save("data/system.xlsx")
    else:
        print "user didn't have this permission"
def check_permission(user,database,action):
    table = load_workbook("data/system.xlsx")['permission']
    db_list = list(iter_cols(table))[0][1:]
    row = db_list.index(database)+2
    action_list = list(iter_rows(table))[0]
    col = action_list.index(action)+1
    allow_user = table.cell(row=row, column=col).value.split(',')
    if user in allow_user:
        return True
    else:
        print "Permission not allowed"
        return False

def check_syntax(sql):
    sql_words = sql.split(' ')
    for i in range(len(sql_words)):
        if sql_words[i] == 'select':
            if sql_words[i+2] == 'from':
                return True
        if sql_words[i] == 'from':
            if sql_words[i+2] == 'where':
                return True
def signup():
    return

def login():
    global user
    print "Please Login:"
    username = raw_input("username: ")
    password = raw_input("password: ")
    if check_login(username,password):
        print "Login Success!Welcome {}! 😊".format(username)
        user = username
    else:
        print "user not exist or password is wrong!😣 Try again."
        login()

def check_login(username,password):
    db = load_workbook("data/system.xlsx")
    #right_pswd = select(password,user,{'username':username})
    table = db['user']
    col_list = list(iter_cols(table))
    try:
        pos = col_list[0].index(username)
    except:
        return False
    right_pswd = col_list[1][pos]
    if hashlib.md5(password).hexdigest() == right_pswd:
        return True
    else:
        return False

def check_Constraint(columns,tablename):    #columns={'a':'xx'}
    db = load_workbook("system/table_information.xlsx")
    table = db[using_dbname]
    rows = []
    rows_list = list(iter_rows(table))  #所有行
    cols_list = list(iter_cols(table))
    for col in columns:
        value = columns[col]
        for i in range(len(cols_list[0])):  #table对应的行
            if cols_list[0][i] == tablename:
                rows.append(i)
        for line in rows:
            if rows_list[line][1] == col:
                typee, is_null, unique, pk, fk = rows_list[line][2:]
                if is_null == '0':
                    if value == '' or value.count(' ')>3:
                        return False
                if unique == '1':
                    if not check_unique(tablename,col,value):
                        return False
                if pk == '1':
                    if not check_unique(tablename,col,value) or value == '':
                        return False
                if '[' in typee:
                    typee, maxlen = re.findall(r'(\w*)\[(\d*)\]', typee) #int[10] => int,10
                else:
                    maxlen = 1000
                if len(value) > maxlen:
                    return False
                if typee == 'int':
                    if type(value) != type(1):
                        return False
                if typee == 'char':
                    if type(value) != type('c'):
                        return False

def check_unique(tablename,column,value):
    table = using_db[tablename]
    col_pos = list(iter_rows(table))[0].index(column)   #第几列
    cols_list = list(iter_cols(table))[col_pos][1:]
    if cols_list.count(value) > 1:  #该列中该值数量
        return False
    else:
        return True


def logout():
    return

def iter_rows(ws):      #表格按行数组形式输出，eg:list(iter_rows(a))
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def iter_cols(ws):      #表格按行数组形式输出，eg:list(iter_rows(a))
    for row in ws.iter_cols():
        yield [cell.value for cell in row]

def query(sql,tag=''):
    sql_word = sql.split(" ")
    if len(sql_word) < 2:
        print "[!] Wrong query!"
        return
    operate = sql_word[0].lower()
    if operate == 'use':
        if sql_word[1] == 'database':
            try:
                use_db(sql_word[2])
            except:
                print "[!]Error"
        else:
            print "[!]Syntax Error.\neg:>use database dbname"
    elif operate == 'create':
        if sql_word[1] == 'database':
            try:
                creat_db(sql_word[2])
            except:
                print "[!]Create Error"
        elif sql_word[1] == 'table':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            print columns_list, using_dbname
            try:
                creat_table(sql_word[2], using_db, using_dbname, columns_list)
            except:
                print "[!]Error"
        elif sql_word[1] == 'view': #creat view test1 as select * from user
            viewname = sql_word[2]
            sql = ' '.join(sql_word[4:])
            view(viewname,sql)

        elif sql_word[1] == 'index':
            return
        else:
            print "[!]Syntax Error."
    elif operate == 'select':
        pos = 0
        for i in range(len(sql_word)):
            if '(' in sql_word[i] and 'select' in sql_word[i]:
                pos = i
        if pos == 3:
            sql2 = sql_word[3][1:-1]
            query(sql2,tag='nesting')
            sql_word[3] = 'tmp'
            sql = ' '.join(sql_word)

        columns = sql_word[1]
        table_name = sql_word[3]
        if len(sql_word) > 4:
            #try:
            limit = sql_word[5].split()
            predicate = 'and'
            symbol = '='
            if ',' in sql_word[5]:
                limit = sql_word[5].split(',')
                predicate = 'and'
            elif '|' in sql_word[5]:
                limit = sql_word[5].split('|')
                predicate = 'or'
            elif '>' in sql_word[5]:
                #limit = sql_word[5].split()
                symbol = '>'
            elif '<' in sql_word[5]:
                #limit = sql_word[5].split()
                symbol = '<'
            elif len(sql_word) > 6:
                if sql_word[6] == 'in':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'in'
                if sql_word[6] == 'like':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'like'
            #except:
                #limit = [].append(sql_word[5])
            #print limit
            for i in range(len(limit)):
                limit[i] = limit[i].split(symbol)
            limit = dict(limit)
            return select(columns, table_name, limit, predicate=predicate, symbol=symbol, tag=tag)
        else:   #没where的情况
            return select(columns, table_name, tag=tag)
    elif operate == 'grant':
        if user != 'admin':
            return  False
        set_permission(sql_word[5], sql_word[3], sql_word[1])
    elif operate == 'revoke':
        if user != 'admin':
            return  False
        del_permission(sql_word[5], sql_word[3], sql_word[1])
    elif operate == 'insert':   #INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
        table_name = sql_word[2]
        """
        #INSERT INTO table_name (select x from xx)
        sql2 = re.findall('\((.*)\)')[0]
        query(sql2,tag='insert')
        """

        columns_list = []
        if '&' in sql:
            cols = sql_word[3].split('&')   #[{xx},{xx}] 多组
            for p in range(len(cols)):
                col = cols[p]
                c = col.split(',')
                for i in range(len(c)):
                    c[i] = c[i].split('=')
                cols[p] = dict(c)
            columns_list = cols
        else:
            cols = sql_word[3].split(',')
            for i in range(len(cols)):
                cols[i] = cols[i].split('=')
            columns_list.append(dict(cols))
        insert(table_name,using_db,using_dbname,columns_list)
    elif operate == 'update':
        return
    elif operate == 'help':
        if sql_word[1] == 'database':
            show_db()
        if sql_word[1] == 'table':
            usdbnm = using_dbname
            use_db('table_information')
            tbname = sql_word[2]
            select('*',usdbnm,{'table':tbname})
        if sql_word[1] == 'view':
            view_name = sql_word[2]
            use_db('view')
            select('sql','sql',{'viewnamw':view_name})
        if sql_word[1] == 'index':
            print "All Index:"
            indexs = os.listdir('data/index/')  # 第二种方法，从保存数据库信息的库中查询
            for index in indexs:
                if '.DS' not in index:
                    print "[*] " + index[:-5]
    else:
        print "[!]Syntax Error."



def run():
    #Initialization()
    welcome()
    login()
    while True:
        command = get_command()
        #print command
        if command == 'quit' or command == 'exit':
            print("[🍻] Thanks for using L-DBMS. Bye~~")
            exit(0)
        elif command == 'help':
            help()
        else:
            query(command)

#####test function
def test_index():
    db = load_workbook("data/system.xlsx")
    index(db,'user','username')

def test_delect():
    db = load_workbook("data/system.xlsx")
    test = {'username': 'aaa'}
    delect('user', db, 'system', test)

def test_create_tb():
    #db = load_workbook("data/system.xlsx")
    use_db('system')
    test = ['id int null pk','user char not_null unique']
    creat_table('test1',using_db,using_dbname,test)

def test_update():
    db = load_workbook("data/system.xlsx")
    test_up = {'username': 'newuser', 'password': 'newpass'}
    test = {'username': 'bbb'}
    update('user', db, 'system', test, test_up)

def test_check():
    user = 'admin'
    db = 'system'
    action = 'select'
    if check_permission(user,db,action):
        print 'ok'
if __name__ == '__main__':
    Initialization()
    run()
